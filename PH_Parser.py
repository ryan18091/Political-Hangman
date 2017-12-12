import datetime as time
import urllib.parse
import urllib.request

import bs4 as bs
from openpyxl import load_workbook

from ParserEmail import *


def runemail():
    print('in runemail')
    wb = load_workbook(filename='ReportSpreadsheet.xlsx')
    WebParserReport = wb['WebParserReport']
    old_data = (WebParserReport['B4']).value

    class OldDataConvert:

        def __init__(self, old_data):
            self.old_data = old_data


        def unicodeTrans(self):
            #translate old_data to remove unwanted characters
            return old_data
            pass

        def tupleConvert(self):
            #convert old_data as tuple to str
            return old_data
            pass


    convertOldData = OldDataConvert(old_data)
    convertOldData.unicodeTrans()
    old_data = str(old_data)
    old_data = old_data.replace(u'\xa0', u' ')
    old_data = old_data.replace(u'\\', u'')
    old_data = old_data.replace(u'xa0', u' ')

    class Politician_Parser:

        def __init__(self, url, query_tag, query_id):
            self.url = url
            self.query_tag = query_tag
            self.query_id = query_id

        def Politician(self):

            try:
                sauce = urllib.request.urlopen(self.url)
                soup = bs.BeautifulSoup(sauce, 'lxml')
                Pol_list = []
                for div in soup.find_all(self.query_tag, class_ = self.query_id):
                    table = str.maketrans(dict.fromkeys('\t\n'))
                    Pol_list.append(div.text.translate(table))
                return Pol_list

            except Exception as e:
                print(e)

    #Instance Variables
    president = Politician_Parser('https://www.whitehouse.gov/', 'span', 'whr-president')
    cabinet = Politician_Parser('https://www.whitehouse.gov/administration/cabinet', 'div', 'field-item even')
    senate_house = Politician_Parser('https://ballotpedia.org/List_of_current_members_of_the_U.S._Congress', 'td', None)

    presData = Politician_Parser.Politician(president)
    cabData = Politician_Parser.Politician(cabinet)
    senHouseData = Politician_Parser.Politician(senate_house)
    data = (str(presData) + str(cabData) + str(senHouseData))
    data = ''.join(data)

    class DataInput_and_Checker:

        def __init__(self, new_data, old_data):
            self.new_data = new_data
            self.old_data = old_data


        def dataDiff(self):
            diff_data = (set(old_data)) - set(new_data)
            return diff_data



    #https://docs.python.org/3/howto/unicode.html
    new_data = (''.join(presData)) + (''.join(cabData)) + (''.join(senHouseData))
    DbInput = DataInput_and_Checker(new_data, old_data)

    #checks for differences in Old vs New Data
    diff = []

    for i in old_data.split():
        if i not in new_data.split():
            diff.append(i)


    if len(diff) == 0:
        diff = 'No Politician Change Found.'



    class xlsxCreator:

        def __init__(self, worksheet, sheet1, sheet2, old_data, new_data, diff_data):
            self.worksheet = worksheet
            self.sheet1 = sheet1
            self.sheet2 = sheet2
            self.old_data = old_data
            self.new_data = new_data
            self.diff_data = diff_data

        def DataClear(self):
            wb = load_workbook(filename=self.worksheet)
            WebParserReport = wb[self.sheet1]
            # UnitTestReport = wb['UnitTestReport']
            WebParserReport['B1'] = str(time.date.today())
            WebParserReport['A4'] = str(self.old_data)
            WebParserReport['B4'] = str(self.new_data)
            WebParserReport['C4'] = str(self.diff_data)
            wb.save('ReportSpreadsheet.xlsx')


        def xlsxDataInput(self):
            wb = load_workbook(filename=self.worksheet)
            WebParserReport = wb[self.sheet1]
            # UnitTestReport = wb['UnitTestReport']
            WebParserReport['B1'] = str(time.date.today())
            WebParserReport['A4'] = str(self.old_data)
            WebParserReport['B4'] = str(self.new_data)
            WebParserReport['C4'] = str(self.diff_data)
            wb.save('ReportSpreadsheet.xlsx')


    xlsxParserCreate = xlsxCreator('ReportSpreadsheet.xlsx', 'WebParserReport', 'UnitTestReport', old_data,
                                   new_data, diff)
    xlsxParserCreate.xlsxDataInput()

    PH_Email()
